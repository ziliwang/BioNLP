#!/usr/bin/env python3
from openpyxl import load_workbook
import re
import os
import queue
import threading
import json

"""
A class using for quering mutation by transvar, and parser transvar results to
VCF mutation format assisted by samtools!
pipeline: extract_mutation_string -> json -> batch_transvar -> ls ->
          parser_transver_output -> json -> samtools -> json -> processdic
"""


def extract_mutation_string(path):
    """
    iterator, for traversal target dir, extract colum 5 in excel(.xlsx), and
    return (file_name, rownum, utation discription)
    """
    for root, dirs, files in os.walk(path):
        for f in files:
            try:
                genename, suffix = f.split('.')
            except:
                continue
            if suffix in ['xlsx']:
                f = os.path.join(root, f)
                wb = load_workbook(filename=f, read_only=True)
                ws = wb[wb.sheetnames[0]]
                for row in ws.rows:
                    if not row[5].value or row[5].row == 1:
                        continue
                    else:
                        reg = re.search(r'(c\.[0-9,_,\*,>,A-Z,a-z,\-,\+]+)',
                                        row[5].value)
                        if reg:
                            yield(genename, row[5].row, reg.groups()[0])
                        else:
                            print('err')


def transvar_bath_query(fl, thread_num=40):
    '''from json file get query str and mutil-prosessing.'''
    f = open(fl, 'r')
    ls = json.loads(f.read())
    f.close()
    myqueue = queue.Queue(0)
    outqueue = queue.Queue(0)
    for i in ls:
        myqueue.put(i)
    threads = []
    for i in range(thread_num):
        thread = threading.Thread(target=transvar_task,
                                  args=(myqueue, outqueue,))
        threads.append(thread)
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join()
    with open('transvar_batch.out', 'w') as f:
        while not outqueue.empty():
            out = outqueue.get()
            for i in out:
                f.write(i + '\n')


def transvar_task(myqueue, outqueue):
    """transvar process unit"""
    while not myqueue.empty():
        prefix, num, text = myqueue.get()
        results = os.popen(
            "transvar canno -i '{0}' --refseq 2>/dev/null| tail -n +2".format(
                                                        prefix + ':' + text))
        lines = [i.strip('\n') for i in results.readlines() if i]
        f_lines = []
        out = []
        for i in lines:  # reqseq results
            if re.search(r'Error', i, re.IGNORECASE) or (
                                    re.search(r'no\_valid', i, re.IGNORECASE)):
                #  filter null result
                pass
            else:
                # store into output list
                f_lines.append(i)
        if f_lines:  # output list has element, and output
            for i in f_lines:
                out.append('{0}\t{1}\t{2}'.format(prefix, num, i))
        else:  # no element in output list, check ensembl and ccds db
            results = os.popen(
                    "transvar canno -i '{0}' --ccds --ensembl 2>/dev/null|"
                    "tail -n +2".format(prefix + ':' + text))
            lines = [i.strip('\n') for i in results.readlines() if i]
            for i in lines:
                if re.search(r'Error', i, re.IGNORECASE) or (
                                    re.search(r'no\_valid', i, re.IGNORECASE)):
                    # filter
                    pass
                else:
                    f_lines.append(i)
                if f_lines:  # output
                    for i in f_lines:
                        out.append('{0}\t{1}\t{2}'.format(prefix, num, i))
        outqueue.put(out)


def parser_transver_output(file):
    '''
    iterator, for extract info from trans output.
    return(filename, rownum, [mutation_type, position, ref_ale, alt_ale, chr],
           raw_transcript_name, mutation_description)
    i. e. ('SLC26A5', '3656', ('snv', '103054757', 'C', 'G', 'chr7'),
          'NM_206885 (protein_coding)', 'chr7:g.103054757C>G/c.293-1198G>C/.')
    '''
    with open(file, 'r') as f:
        raw = [i.strip('\n') for i in f.readlines() if i]
    dic = {}
    for line in raw:
        prefix, row, input_str, trans, gene, strand, position, info1, info2 =\
            line.split('\t')
        try:
            dic[prefix][row].append((input_str, trans, gene, strand, position,
                                     info1, info2))
        except:
            try:
                dic[prefix][row] = [(input_str, trans, gene, strand, position,
                                     info1, info2)]
            except:
                dic[prefix] = {}
                dic[prefix][row] = [(input_str, trans, gene, strand, position,
                                     info1, info2)]
    for f in dic.keys():
        for r in dic[f].keys():
            for input_str, trans, gene, strand, position, info1, info2 in \
                                                                    dic[f][r]:
                yield(f, r, mutation_reg(position), trans, position)


def mutation_reg(position):
    '''
    classify mutation into snv, insert, dup, del-short, del-long, delins,
    unkown, and return vcf format.
    '''
    mutation_de = position.split('/')[0]
    chr = mutation_de.split(':')[0]
    # sn
    m = re.search(r'(\d+)([ATCG])>([ATCG])', mutation_de)
    if m:
        return('snv', m.group(1), m.group(2), m.group(3), chr)
    # insert  trans [2]->ref [3]+=ref
    m = re.search(r'(\d+)_\d+ins([AGCT]+)', mutation_de)
    if m:
        return('insert', m.group(1), '-', m.group(2), chr)
    # dup
    m = re.search(r'(\d+)_?\d*dup([ATCG]+)', mutation_de)
    if m:
        return('dup', m.group(1), m.group(2), m.group(2)*2, chr)
    # delshort
    m = re.search(r'(\d+)_?\d*del([ATCG]+)', mutation_de)
    if m:
        return('del-short', m.group(1), m.group(2), '-', chr)
    # del long > trans  [2] - > ref[[1]:[2]+[1]-1]
    m = re.search(r'(\d+)_\d+del([0-9]+)', mutation_de)
    if m:
        return('del-long', m.group(1), m.group(2), '-', chr)
    # delins [2] = ref[[1],[2]]
    m = re.search(r'(\d+)_(\d+)delins([ATCG]+)', mutation_de)
    if m:
        return('delins', m.group(1), m.group(2), m.group(3), chr)
    # undected
    return('unkown', 'un', 'un', 'un', chr)


def samtools_check(fl):
    f = open(fl, 'r')
    data = json.loads(f.read())
    f.close()
    for gene, r, mutation_reg, trans, position in data:
        if mutation_reg[0] == 'insert':
            mutation_reg[2] = query(mutation_reg[1], mutation_reg[1],
                                    mutation_reg[4])
            mutation_reg[3] = mutation_reg[2] + mutation_reg[3]
        elif mutation_reg[0] == 'del-long':
            mutation_reg[1] = int(mutation_reg[1]) - 1
            mutation_reg[3] = query(mutation_reg[1], mutation_reg[1],
                                    mutation_reg[4])
            mutation_reg[2] = query(
                                mutation_reg[1],
                                int(mutation_reg[1])+int(mutation_reg[2]),
                                mutation_reg[4])
        elif mutation_reg[0] == 'delins':
            mutation_reg[2] = query(mutation_reg[1], mutation_reg[2],
                                    mutation_reg[4])
        elif mutation_reg[0] == 'del-short':
            mutation_reg[1] = int(mutation_reg[1]) - 1
            mutation_reg[3] = query(mutation_reg[1], mutation_reg[1],
                                    mutation_reg[4])
            mutation_reg[2] = mutation_reg[3] + mutation_reg[2]
        else:
            pass
    out = open('transed.json', 'w')
    json.dump(data, out)
    out.close()


def query(start, end, chr):
    return(''.join([i.strip('\n') for i in os.popen(
           'samtools faidx /home/zhaok/Data/bundle/hg19/ucsc.hg19.fasta'
           ' {2}:{0}-{1}'.format(start, end, chr)).readlines()
           if not re.search('>', i)]).upper())


def trans_to_dic(ls):
    '''trans samtools fixed output into dict'''
    dic = {}
    for f, r, mutation_reg, trans, pos in ls:
        try:
            dic[f][r].append((mutation_reg, trans, pos))
        except:
            try:
                dic[f][r] = [(mutation_reg, trans, pos)]
            except:
                dic[f] = {}
                dic[f][r] = [(mutation_reg, trans, pos)]
    return dic


def processdic(dic, path):
    """main logic process"""
    for root, dirs, files in os.walk(path):
        for f in files:
            try:
                genename, suffix = f.split('.')
            except:
                continue
            if suffix in ['xlsx'] and genename in dic.keys():
                f = os.path.join(root, f)
                wb = load_workbook(filename=f, read_only=True)
                with open(genename + '.txt', 'w') as f:
                    ws = wb[wb.sheetnames[0]]
                    for row in ws.rows:
                        if row[5].value:
                            if row[5].row == 1:
                                continue
                            logic_part(row, dic, genename, f)


def logic_part(row, dic, genename, f):
    check = []
    if str(row[5].row) in dic[genename].keys():
        for ann, trans, mix in dic[genename][str(row[5].row)]:
            ty, pos, ref, mut, ch = ann
            transname = trans.split(' ')[0]
            HGVS = transname + ':' + mix.split('/')[1]
            check.append([transname, ch, pos, ref, mut, HGVS])
        check_sw = False
        for i in check:
            if re.search(i[0], row[5].value):
                check_sw = True
                try:
                    if (re.sub('chr', '', i[1], flags=re.IGNORECASE) ==
                            re.sub('chr', '', str(row[1].value),
                                   flags=re.IGNORECASE)) and (
                            int(i[2]) == int(row[2].value)) and (
                                i[3] == row[3].value) and (
                                  i[4] == row[4].value):
                        writeh([genename, 'PASS'] + i[1:] +
                               [i.value for i in row[1:31]], f)
                    else:
                        writeh([genename, 'FAILED'] + i[1:] +
                               [i.value for i in row[1:31]], f)
                except:
                    writeh([genename, 'NoData'] + i[1:] +
                           [i.value for i in row[1:31]], f)
        if not check_sw:
            i = check[0]
            writeh([genename, 'Invalid'] + i[1:] +
                   [i.value for i in row[1:31]], f)
    else:
        writeh([genename, 'NoData'] + ['None']*5 +
               [i.value for i in row[1:31]], f)


def writeh(ls, f):
    ls = [str(i) if i else 'Null' for i in ls]
    ls = [re.sub('\t', '\s', i) for i in ls]
    f.write('\t'.join(ls).replace('\n', ' ') + '\n')


def vcflization(txtdir):
    '''change all mutation description into a vcf-support format, and I called
    'vcflization'. This function batch process the txt format input, and output
    same txt file with vcflization and recompare the position, chr, ref, ale
    info.'''
    for fn in os.listdir(txtdir):
        if os.path.isfile(os.path.join(txtdir, fn)):
            if re.search('txt$', fn):
                if not os.path.exists(os.path.join(txtdir, 'vcflization')):
                    os.makedirs(os.path.join(txtdir, 'vcflization'))
                with open(os.path.join(txtdir, fn), 'r') as infh:
                    with open(os.path.join(txtdir, 'vcflization', fn), 'w'
                              ) as outfh:
                        for line in [i.split('\t') for i in
                                     infh.read().split('\n') if i]:
                            line = line[:7] + \
                                   normalization(line[2], line[7:11]) + \
                                   line[7:]
                            line[1] = recompare(line[1], line[7:11], line[2:6])
                            outfh.write('\t'.join(line) + '\n')


def normalization(ichr, ls):
    """convert format into vcf-support format, host a reference chr which
    is used to point the chr when chr info unavail in origin.
    input: ref_chr, (chr, pos, ref, ale)
    output: (chr, pos, ref, ale)
    """
    chr, pos, ref, ale = ls
    chr_reg = re.search('([0-9,M,m]+)', chr)
    if chr_reg:
        chr = 'chr' + chr_reg.group(1)
    else:
        chr = ichr
    pos_reg = re.search('(\d+)', str(pos))
    if pos_reg:
        if ref == '-':  # insert
            if len(pos_reg.groups()) == 2:
                before_pos = pos_reg.group(1)
                before = query(before_pos, before_pos, chr)
                return([chr, before_pos, before, before + ale])
            else:
                # default insert after pos
                before_pos = pos_reg.group(1)
                before = query(before_pos, before_pos, chr)
                return([chr, before_pos, before, before + ale])
        elif ale == '-':  # default pos is the start del site
            before_pos = str(int(pos_reg.group(1)) - 1)
            before = query(before_pos, before_pos, chr)
            return([chr, before_pos, before + ref, before])
        else:
            return([chr, pos, ref, ale])
    else:
        return([chr, pos, ref, ale])


def recompare(status, ls, ils):
    '''compare the two mutation info with previous kown status. When status is
    'FAILED', start comparing and re-compute status, otherwise, return origin
    status
    '''
    chr, pos, ref, ale = ls
    ichr, ipos, iref, iale = ils
    if chr == ichr and str(pos) == str(ipos) and ref == iref and (
                ale == iale) and status == 'FAILED':
        return 'PASS'
    return status
